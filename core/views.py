from django.shortcuts import render,redirect
from django.contrib.sites.shortcuts import get_current_site
from django.http import HttpResponse
import openpyxl
from .forms import RegisterForm, LoginForm, CustomSetPasswordForm
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.models import User
from django.contrib import messages

from django.core.mail import send_mail
from django.utils.http import urlsafe_base64_encode, urlsafe_base64_decode
from django.template.loader import render_to_string
from django.utils.encoding import force_bytes
from django.contrib.auth.tokens import default_token_generator

from django.core.mail import send_mail
from django.conf import settings

# Create your views here.
def signin(request):
    if request.method == 'POST':
        form = LoginForm(request.POST)
        if form.is_valid():
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)
                if user.is_staff:
                    return redirect('manager')
                return redirect('home')
            else:
                form.add_error(None, "Invalid email or password")
    else:
        form = LoginForm()


    return render(request, 'core/signin.html', {'form': form})


def signup(request):
    if request.method == 'POST':
        form = RegisterForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            if 'happyholiday' in user.email:
                user.is_staff = True
            user.save()
            
            username = form.cleaned_data['username']
            password = form.cleaned_data['password1']
            user = authenticate(username=username, password=password)
            if user:
                login(request, user)
                if 'happyholiday' in user.email:
                    return redirect('manager')
                return redirect('home')
        else:
            print('Form errors:', form.errors)
    else:
        form = RegisterForm()
    return render(request, 'core/signup.html', {'form': form})


def logout_(request):
    logout(request)
    return redirect('home')

def forgot_password(request):
    if request.method == 'POST':
        email = request.POST.get('email')
        associated_user = User.objects.filter(email=email).first()
        print(f"Email entered: {email}")
        print(f"Associated user found: {associated_user}")

        if associated_user:
            subject = 'Password Reset Requested'
            email_template_name = 'core/password_reset_email.html'
            context = {
                'email': associated_user.email,
                'domain': get_current_site(request).domain,
                'site_name': "Happy Holiday",
                'uid': urlsafe_base64_encode(force_bytes(associated_user.pk)),
                'token': default_token_generator.make_token(associated_user),
                'protocol': 'https' if request.is_secure() else 'http',
            }

            email_content = render_to_string(email_template_name, context)
            send_mail(subject, email_content, '200107017@stu.sdu.edu.kz', [associated_user.email], fail_silently=False)
            messages.success(request, 'A reset code has been sent to your email address.')
        else:
            messages.error(request, 'The specified email is not registered.')

    return render(request, 'core/forgot-password.html', {})


def password_reset_success(request):
    return render(request, 'core/password_reset_success.html', {})

def change_password(request, uidb64, token):
    try:
        uid = urlsafe_base64_decode(uidb64).decode()
        user = User.objects.get(pk = uid)
        print(uid, user)
    except (TypeError, ValueError, OverflowError, User.DoesNotExist):
        print(urlsafe_base64_encode(uidb64).decode())
        user = None

    if user and default_token_generator.check_token(user, token):
        if request.method == 'POST':
            form = CustomSetPasswordForm(user, request.POST)
            
            if form.is_valid():
                form.save()
                return redirect('home')
        else:
            form = CustomSetPasswordForm(user)
        return render(request, 'core/password_reset.html', {'form': form, 'correct': True})
    else:
        return render(request, 'core/password_reset.html', {'correct': False})



def home(request):
    user_id = request.user.id

    return render(request, 'core/home.html', {'id': user_id})

def profile(request, id):
    user = User.objects.get(pk=id)

    return render(request, 'core/profile.html', {'user': user, 'id': id})



def manager_page(request):


    return render(request, 'core/manager_home.html', {})

def location(request):

    return render(request, 'core/locations.html', {})

def suppliers(request):
    return render(request, 'core/suppliers.html')

def about_us(request):
    return render(request, 'core/about_us.html')  # New view for About Us

def contact(request):
    return render(request, 'core/contact.html')


# suppliers
def photographers(request):
    return render(request, 'core/suppliers/photographers.html')

def decorators(request):
    return render(request, 'core/suppliers/decorators.html')

def menu_bar(request):
    return render(request, 'core/suppliers/menu_bar.html')

def choreographers(request):
    return render(request, 'core/suppliers/choreographers.html')

def designers(request):
    return render(request, 'core/suppliers/designers.html')

def venue_planners(request):
    return render(request, 'core/suppliers/venue_planners.html')

def makeup_artists(request):
    return render(request, 'core/suppliers/makeup_artists.html')



def generate_report(request):
    # Create a new Excel file
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Report"

    # Fill with sample data
    ws['A1'] = "ID"
    ws['B1'] = "Name"
    ws.append([1, 'John Doe'])
    ws.append([2, 'Jane Smith'])

    # Create a response with Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
    wb.save(response)
    return response